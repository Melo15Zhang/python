import random
from openpyxl import Workbook
import datetime

upper = ["A", "B", "C", "D", "E", "F", "G",
         "H", "I", "J", "K", "L", "M", "N",
         "O", "P", "Q", "R", "S", "T",
         "U", "V", "W", "X", "Y", "Z"]
lower = ["a", "b", "c", "d", "e", "f", "g",
         "h", "i", "j", "k", "l", "m", "n",
         "o", "p", "q", "r", "s", "t",
         "u", "v", "w", "x", "y", "z"]
digital = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
special = ["!", "@", "#", "$", "%"]
map = {"1": upper, "2": lower, "3": digital, "4": special}

wb = Workbook()
ws = wb.active  # 激活 worksheet
ws['A1'] = "密码"
for i in range(5000):
    str = ""
    str += map.get("1")[random.randrange(0, len(upper))]
    str += map.get("2")[random.randrange(0, len(lower))]
    str += map.get("3")[random.randrange(0, len(digital))]
    str += map.get("4")[random.randrange(0, len(special))]
    for j in range(4):
        num = random.randrange(1, 100)
        select = (num % 4 + 1)
        list = map.get("{}".format(select))
        if select == 3:
            str += list[random.randrange(0, len(digital))]
        elif select == 4:
            str += list[random.randrange(0, len(special))]
        else:
            str += list[random.randrange(0, len(upper))]
    ce = 'A{}'.format(i + 2)  # 数据可以直接分配到单元格中
    ws[ce] = str
# Python 类型会被自动转换
ws['A5002'] = datetime.datetime.now().strftime("%Y-%m-%d")
wb.save("pass.xlsx")  # 保存文件
wb.close()

